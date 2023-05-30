from cosmospy import generate_wallet
from openpyxl import Workbook

#Ввод числа
kolvo = input('Сколько? ')

while True:
	try:
		kolvo = int(kolvo)
	except:
		print('ТЫ уродец, введи число, еблан')
	if isinstance(kolvo,int):
		break
	else:
		kolvo = input('Сколько? ')

#Создание таблицы
wb = Workbook()
ws = wb.active
ws['A1'] = 'address'
ws['B1'] = 'seed'

#Создание кошелька
def wallet():
	wallet = generate_wallet()
	atas = []
	atas.append(wallet['address'])
	atas.append(wallet['seed'])
	return atas


for a in range(kolvo):
	ws.append(wallet())



wb.save('cosmos.xlsx')
